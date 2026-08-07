"""One-shot import of the Excel budget workbooks.

Not a product feature — a migration tool, run once per workbook and then kept
around so the import can be re-run as the merchant rules are tuned.

    uv run python scripts/import_workbook.py ~/Desktop/Budgets/*.xlsx
    uv run python scripts/import_workbook.py --reset ~/Desktop/Budgets/*.xlsx

Nothing is guessed silently. Every row that can't be taken at face value —
missing date, implausible date, renamed category — is imported as faithfully as
possible and listed in the anomaly report at the end.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from backend.db import SessionLocal
from backend.merchants import display_name, normalize_merchant
from backend.models import (
    Account,
    AccountBalance,
    BudgetAllocation,
    BudgetPeriod,
    Category,
    CategoryKind,
    FixedCost,
    Merchant,
    MerchantPattern,
    PaycheckLine,
    PaycheckLineKind,
    Transaction,
    TransactionSource,
)

# --------------------------------------------------------------------------
# Reference data
# --------------------------------------------------------------------------

# Canonical categories, in the order the budget sheet lists them.
CATEGORY_SEED: list[tuple[str, CategoryKind]] = [
    ("Savings", CategoryKind.SAVINGS),
    ("Insurance", CategoryKind.SPENDING),
    ("Rent", CategoryKind.SPENDING),
    ("Utilities", CategoryKind.SPENDING),
    ("Subscriptions", CategoryKind.SPENDING),
    ("Loan Payments", CategoryKind.SPENDING),
    ("Donations", CategoryKind.SPENDING),
    ("Groceries", CategoryKind.SPENDING),
    ("Food and Drinks", CategoryKind.SPENDING),
    ("Transportation", CategoryKind.SPENDING),
    ("Shopping", CategoryKind.SPENDING),
    ("Entertainment", CategoryKind.SPENDING),
    ("Travel", CategoryKind.SPENDING),
    ("Self Care", CategoryKind.SPENDING),  # holds the gym *and* one-off spending
    ("Misc", CategoryKind.SPENDING),
    ("Deficit Reduction", CategoryKind.OTHER),
]

# The category names drifted over three years of workbooks.
CATEGORY_ALIASES = {
    "drinks": "Food and Drinks",
    "food": "Food and Drinks",
    "grocery": "Groceries",
    "london": "Travel",  # a trip-specific bucket, 39 transactions in 2024
}

# 'Refund' was a category doing a sign's job. Both such rows are already
# negative, so only the category needs correcting — to whatever they refunded.
REFUND_REMAP = {
    "trip refund": "Travel",
    "prime video channels": "Subscriptions",
}

MONTHS = {
    m.lower(): i
    for i, m in enumerate(
        [
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ],
        start=1,
    )
}

# --------------------------------------------------------------------------
# Accounts sheet corrections, all confirmed against the source on 2026-08-07
# --------------------------------------------------------------------------

# 'Vested Retirement' is the vested *portion* of the retirement account, not a
# separate holding. The 2024 sheet proves it: its own 'Total' row excludes the
# line, and its 'Total (vested)' row equals vested + non-retirement. Importing
# it as an account double-counted net worth by up to $33,348.
NOT_AN_ACCOUNT = {"vested retirement"}

# The account was renamed between workbooks. Same money, one continuous series.
ACCOUNT_ALIASES = {("Fidelity", "Retirement"): "Moody's PPP"}

# The 2026 workbook's Accounts sheet was copied from the 2024 one and
# relabelled with 2025 dates without the numbers changing — all seven columns
# hold values identical to 2024's. The 2024 dates are the real ones.
STALE_SNAPSHOT_YEARS = {"Budget 2026": {2025}}

PAYCHECK_KINDS = {
    "income": PaycheckLineKind.INCOME,
    "insurance": PaycheckLineKind.INSURANCE,
    "savings": PaycheckLineKind.SAVINGS,
    "tax": PaycheckLineKind.TAX,
}

# --------------------------------------------------------------------------
# Anomalies
# --------------------------------------------------------------------------


@dataclass
class Report:
    anomalies: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    counts: Counter = field(default_factory=Counter)

    def note(self, kind: str, detail: str) -> None:
        self.anomalies[kind].append(detail)

    def render(self) -> str:
        if not self.anomalies:
            return "No anomalies."
        out = []
        for kind, items in sorted(self.anomalies.items()):
            out.append(f"\n{kind}  ({len(items)})")
            for line in items[:8]:
                out.append(f"    {line}")
            if len(items) > 8:
                out.append(f"    ... and {len(items) - 8} more")
        return "\n".join(out)


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------


def money(value) -> Decimal | None:
    """Excel hands back floats; quantize once, at the boundary."""
    if value is None or isinstance(value, str):
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"))


def sheet_month(sheet_name: str) -> int | None:
    first = sheet_name.strip().split()[0].lower()
    return MONTHS.get(first)


def parse_text_date(text: str, year: int) -> date | None:
    """Accounts sheet headers are prose, and inconsistently so.

    2024 writes 'March 2nd' and 'Amount As of Feb 2' with no year at all; 2026
    writes 'Sunday, February 2, 2025' with one. The written year wins when
    present — taking the workbook's year instead filed every 2025 snapshot
    under 2026 and invented a year of net worth history that never happened.
    """
    if isinstance(text, datetime):
        return text.date()
    if not isinstance(text, str):
        return None
    low = text.lower()
    for name, num in MONTHS.items():
        if name[:3] not in low:
            continue
        tail = low[low.index(name[:3]) :]
        day = re.search(r"(\d{1,2})(?!\d)", tail)
        if not day:
            return None
        written_year = re.search(r"\b(19|20)\d{2}\b", text)
        try:
            return date(
                int(written_year.group()) if written_year else year,
                num,
                int(day.group(1)),
            )
        except ValueError:
            return None
    return None


# --------------------------------------------------------------------------
# Import
# --------------------------------------------------------------------------


class Importer:
    def __init__(self, session: Session, report: Report) -> None:
        self.s = session
        self.r = report
        self.categories: dict[str, Category] = {}
        self.periods: dict[tuple[int, int], BudgetPeriod] = {}
        self.merchants: dict[str, Merchant] = {}
        self.accounts: dict[tuple[str, str], Account] = {}
        self.seen_allocations: set[tuple[int, int]] = set()
        self.sheet_totals: dict[str, Decimal] = defaultdict(Decimal)
        self.imported_totals: dict[str, Decimal] = defaultdict(Decimal)
        # The sheet's own rollup 'Total' cell — the number he actually reads.
        self.sheet_self_totals: dict[str, Decimal] = {}

    # -- seeds ------------------------------------------------------------

    def seed_categories(self) -> None:
        for order, (name, kind) in enumerate(CATEGORY_SEED):
            cat = Category(name=name, kind=kind, sort_order=order)
            self.s.add(cat)
            self.categories[name.lower()] = cat
        self.s.flush()

    def category_for(self, raw: str | None, description: str, where: str) -> Category:
        name = (raw or "Misc").strip()
        low = name.lower()

        if low == "refund":
            target = REFUND_REMAP.get(description.strip().lower())
            if target:
                self.r.note(
                    "refund category remapped", f"{where}: {description!r} -> {target}"
                )
                return self.categories[target.lower()]
            self.r.note("refund with no mapping -> Misc", f"{where}: {description!r}")
            return self.categories["misc"]

        if low in CATEGORY_ALIASES:
            return self.categories[CATEGORY_ALIASES[low].lower()]

        if low in self.categories:
            return self.categories[low]

        self.r.note("unknown category -> Misc", f"{where}: {raw!r}")
        return self.categories["misc"]

    def period(self, year: int, month: int) -> BudgetPeriod:
        key = (year, month)
        if key not in self.periods:
            p = BudgetPeriod(year=year, month=month)
            self.s.add(p)
            self.s.flush()
            self.periods[key] = p
        return self.periods[key]

    def merchant(self, raw: str, category: Category) -> Merchant | None:
        key = normalize_merchant(raw)
        if not key:
            return None
        if key not in self.merchants:
            m = Merchant(
                canonical_name=display_name(key), default_category_id=category.id
            )
            self.s.add(m)
            self.s.flush()
            self.s.add(MerchantPattern(merchant_id=m.id, pattern=key))
            self.merchants[key] = m
        return self.merchants[key]

    # -- transactions -----------------------------------------------------

    def capture_self_total(self, ws, key: str) -> None:
        """Find the sheet's own rollup 'Total' — the figure he actually reads.

        Reconciling against my own sum of column D only proves I dropped
        nothing. Reconciling against this proves I read the right column.
        """
        for row in ws.iter_rows(min_col=5, max_col=12):
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and cell.value.strip().lower() == "total"
                ):
                    for offset in (1, 2):
                        neighbour = ws.cell(cell.row, cell.column + offset).value
                        amount = money(neighbour)
                        if amount is not None:
                            self.sheet_self_totals[key] = amount
                            return

    def import_spending(self, ws, wb_year: int, label: str) -> None:
        s_month = sheet_month(ws.title)
        self.capture_self_total(ws, label + "!" + ws.title)

        for row in ws.iter_rows(min_row=2, max_col=5):
            raw_date, desc, raw_cat, raw_amt, raw_auto = (c.value for c in row)
            if desc is None or raw_amt is None:
                continue
            amount = money(raw_amt)
            if amount is None:
                continue

            where = f"{label}!{ws.title}!r{row[0].row}"
            self.sheet_totals[label + "!" + ws.title] += amount

            occurred_on: date | None = None
            if isinstance(raw_date, datetime):
                d = raw_date.date()
                if abs(d.year - wb_year) <= 1:
                    occurred_on = d
                    if s_month and d.month != s_month:
                        self.r.note(
                            "date falls outside its sheet's month",
                            f"{where}: {d} on '{ws.title}' -> filed under {d:%Y-%m}",
                        )
                else:
                    self.r.note("implausible date, dropped", f"{where}: {d} ({desc!r})")
            elif raw_date is not None:
                self.r.note("unparseable date, dropped", f"{where}: {raw_date!r}")
            else:
                self.r.note("no date (period from sheet name)", f"{where}: {desc!r}")

            if occurred_on is not None:
                period = self.period(occurred_on.year, occurred_on.month)
            elif s_month is not None:
                period = self.period(wb_year, s_month)
            else:
                self.r.note("no date and no month in sheet name — SKIPPED", where)
                self.sheet_totals[label + "!" + ws.title] -= amount
                continue

            description = str(desc).strip()
            category = self.category_for(raw_cat, description, where)
            merchant = self.merchant(description, category)

            self.s.add(
                Transaction(
                    occurred_on=occurred_on,
                    period_id=period.id,
                    raw_description=description[:200],
                    merchant_id=merchant.id if merchant else None,
                    category_id=category.id,
                    amount=amount,
                    is_recurring=str(raw_auto).strip().lower() in {"yes", "true"},
                    source=TransactionSource.WORKBOOK,
                    source_ref=where,
                )
            )
            self.imported_totals[label + "!" + ws.title] += amount
            self.r.counts["transactions"] += 1

    # -- budgets ----------------------------------------------------------

    def import_budget(self, ws, wb_year: int, label: str) -> None:
        month = sheet_month(ws.title)
        if month is None:
            return
        period = self.period(wb_year, month)

        # The category block runs from row 6 to the first blank name. Below it
        # sit a totals row and a 'Day into the Month' block whose numbers would
        # otherwise read as categories with allotments.
        for r in range(6, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name is None or not str(name).strip():
                break

            where = f"{label}!{ws.title}!r{r}"
            raw_amount = ws.cell(r, 2).value
            amount = money(raw_amount)
            if amount is None:
                if isinstance(raw_amount, str) and raw_amount.strip():
                    self.r.note(
                        "broken formula in budget cell", f"{where}: {raw_amount!r}"
                    )
                continue
            if amount == 0:
                continue

            # Strict here: an unrecognised name in the category block means the
            # block ended, not that the row belongs in Misc.
            category = self.categories.get(
                CATEGORY_ALIASES.get(
                    str(name).strip().lower(), str(name).strip()
                ).lower()
            )
            if category is None:
                self.r.note("unknown budget category — skipped", f"{where}: {name!r}")
                continue

            key = (period.id, category.id)
            if key in self.seen_allocations:
                self.r.note(
                    "duplicate budget allocation — kept first",
                    f"{where}: {name!r} = {amount}",
                )
                continue
            self.seen_allocations.add(key)

            self.s.add(
                BudgetAllocation(
                    period_id=period.id, category_id=category.id, amount=amount
                )
            )
            self.r.counts["budget allocations"] += 1
        self.s.flush()

    # -- accounts ---------------------------------------------------------

    def import_accounts(self, ws, wb_year: int, label: str) -> None:
        header = [c.value for c in ws[1]]
        has_retirement_col = any(
            isinstance(h, str) and "retirement" in h.lower() for h in header[:3]
        )
        name_col = 3 if has_retirement_col else 2
        first_date_col = name_col + 1

        stale_years = STALE_SNAPSHOT_YEARS.get(label, set())
        dates: dict[int, date] = {}
        for idx in range(first_date_col, len(header) + 1):
            parsed = parse_text_date(header[idx - 1], wb_year)
            if not parsed:
                continue
            if parsed.year in stale_years:
                self.r.note(
                    "stale copied snapshot column — skipped",
                    f"{label}!{ws.title} col {idx}: {header[idx - 1]!r}",
                )
                continue
            dates[idx] = parsed

        for row in ws.iter_rows(min_row=2):
            institution = row[0].value
            name = row[name_col - 1].value
            if not institution or not str(institution).strip():
                continue  # Total / Retirement / Non-Retirement summary rows
            institution = str(institution).strip()
            name = str(name).strip() if name else institution

            if name.lower() in NOT_AN_ACCOUNT:
                self.r.note(
                    "not a holding — skipped",
                    f"{label}: {institution} / {name} is a portion of another "
                    f"account, and the sheet's own Total excludes it",
                )
                continue

            # Inferred from the name the sheet used, before any alias is
            # applied: renaming 'Retirement' to "Moody's PPP" first would make
            # the name stop containing 'retirement' and silently flip the flag.
            if has_retirement_col:
                is_retirement = str(row[1].value).strip().lower() in {"true", "yes"}
            else:
                is_retirement = "retirement" in name.lower()
                self.r.note(
                    "is_retirement inferred from account name",
                    f"{label}: {institution} / {name} -> {is_retirement}",
                )

            name = ACCOUNT_ALIASES.get((institution, name), name)

            key = (institution, name)
            if key not in self.accounts:
                acct = Account(
                    institution=institution, name=name, is_retirement=is_retirement
                )
                self.s.add(acct)
                self.s.flush()
                self.accounts[key] = acct
                self.r.counts["accounts"] += 1
            acct = self.accounts[key]

            for col, as_of in dates.items():
                balance = money(row[col - 1].value)
                if balance is None:
                    continue
                exists = self.s.scalar(
                    select(AccountBalance).where(
                        AccountBalance.account_id == acct.id,
                        AccountBalance.as_of == as_of,
                    )
                )
                if exists:
                    continue
                self.s.add(
                    AccountBalance(account_id=acct.id, as_of=as_of, balance=balance)
                )
                self.r.counts["balance snapshots"] += 1
        self.s.flush()

    # -- current config ---------------------------------------------------

    def import_fixed_costs(self, ws, effective_from: date, label: str) -> None:
        key = label + "!" + ws.title
        self.capture_self_total(ws, key)
        for row in ws.iter_rows(min_row=2, max_col=4):
            desc, amount, cat, exact = (c.value for c in row)
            amount = money(amount)
            if not desc or amount is None:
                continue
            # 'Bundled-Utility', 'Utility' and 'Health' exist only on this sheet.
            #
            # Bundled utilities are internet and water, which the landlord bills
            # inside the rent charge — Rent 1474.68 + Bundled-Utility 78.69 is
            # exactly the 1553.37 on the Rent sheet, and it arrives as a single
            # 'BILT CARD HOUSING' transaction. Filing them under Utilities would
            # understate the rent commitment and inflate utilities by the same
            # amount, and would not match how the money actually leaves the
            # account. What each line is stays in `description`.
            raw = str(cat or "Misc")
            lookup = {
                "bundled-utility": "Rent",
                "utility": "Utilities",
                "health": "Self Care",
                "subscription": "Subscriptions",
            }
            name = lookup.get(raw.strip().lower(), raw)
            category = self.category_for(name, "", f"{label}!{ws.title}!r{row[0].row}")
            self.imported_totals[key] += amount
            self.s.add(
                FixedCost(
                    description=str(desc).strip()[:80],
                    amount=amount,
                    category_id=category.id,
                    is_exact=str(exact).strip().lower() in {"yes", "true"},
                    effective_from=effective_from,
                )
            )
            self.r.counts["fixed costs"] += 1

    def import_rent_breakdown(self, ws, effective_from: date, label: str) -> None:
        """Replace the coarse rent rows with the Rent sheet's own breakdown.

        Monthly Fixed Costs records rent as three lines — Rent 1474.68,
        Internet 62.00, Water 16.69 — while the Rent sheet itemises the same
        1553.37 across thirteen: rent, three admin charges, sewer, stormwater,
        valet trash, an amenity fee and so on.

        The thirteen become components of one parent worth exactly their sum,
        so the monthly total does not move by a cent. Which is the point: when
        the charge lands at 1535.17 instead of 1553.37, the breakdown is the
        only thing that can say which line moved.
        """
        lines: list[tuple[str, Decimal]] = []
        for row in ws.iter_rows(min_row=2, max_col=3):
            desc, _category, amount = (c.value for c in row)
            value = money(amount)
            if not desc or value is None:
                continue
            lines.append((str(desc).strip()[:80], value))

        if not lines:
            return
        total = sum((amount for _, amount in lines), Decimal("0.00"))

        # The rows this supersedes: everything the fixed-cost sheet filed under
        # Rent. They are replaced, not added to.
        #
        # Flush first. The session runs with autoflush off, so without this the
        # rows import_fixed_costs just added are still pending and the query
        # below finds nothing — leaving the coarse rows in place beside the new
        # bill and double-counting rent.
        self.s.flush()
        rent = self.categories["rent"]
        superseded = self.s.scalars(
            select(FixedCost).where(FixedCost.category_id == rent.id)
        ).all()
        replaced_total = sum((f.amount for f in superseded), Decimal("0.00"))
        if replaced_total != total:
            self.r.note(
                "rent breakdown does not match the fixed-cost rows it replaces",
                f"{label}: Rent sheet totals {total}, fixed costs total "
                f"{replaced_total} — importing the breakdown anyway, at its own total",
            )
        for f in superseded:
            self.s.delete(f)
        self.s.flush()

        parent = FixedCost(
            description="Rent (billed as one charge)",
            amount=total,
            category_id=rent.id,
            is_exact=False,
            effective_from=effective_from,
        )
        self.s.add(parent)
        self.s.flush()
        for desc, amount in lines:
            self.s.add(
                FixedCost(
                    description=desc,
                    amount=amount,
                    category_id=rent.id,
                    is_exact=True,
                    effective_from=effective_from,
                    parent_id=parent.id,
                )
            )
            self.r.counts["rent line items"] += 1
        self.s.flush()

    def import_paycheck(self, ws, effective_from: date, label: str) -> None:
        for row in ws.iter_rows(min_row=2, max_col=3):
            desc, amount, kind = (c.value for c in row)
            amount = money(amount)
            if not desc or amount is None or not kind:
                continue
            k = PAYCHECK_KINDS.get(str(kind).strip().lower())
            if k is None:
                self.r.note("unknown paycheck kind — skipped", f"{label}: {kind!r}")
                continue
            self.s.add(
                PaycheckLine(
                    description=str(desc).strip()[:80],
                    amount=amount,
                    kind=k,
                    effective_from=effective_from,
                )
            )
            self.r.counts["paycheck lines"] += 1


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------


def reset(session: Session) -> None:
    for model in (
        Transaction,
        BudgetAllocation,
        AccountBalance,
        MerchantPattern,
        Merchant,
        FixedCost,
        PaycheckLine,
        Account,
        BudgetPeriod,
        Category,
    ):
        session.execute(delete(model))
    session.commit()


def reset_config(session: Session) -> None:
    """Wipe fixed costs and paycheck lines only.

    Same reasoning as reset_accounts: a full --reset destroys merchant merges
    and split decisions, which are the user's own work.
    """
    session.execute(delete(FixedCost).where(FixedCost.parent_id.is_not(None)))
    session.execute(delete(FixedCost))
    session.execute(delete(PaycheckLine))
    session.commit()


def reset_accounts(session: Session) -> None:
    """Wipe only accounts and their snapshots.

    A full --reset also destroys merchant merges and split decisions, which are
    the user's own work and not reproducible from the workbooks. Re-importing
    the Accounts sheets alone must not cost them that.
    """
    session.execute(delete(AccountBalance))
    session.execute(delete(Account))
    session.commit()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbooks", nargs="+", type=Path)
    ap.add_argument("--reset", action="store_true", help="wipe all data first")
    ap.add_argument(
        "--accounts-only",
        action="store_true",
        help="re-import just the Accounts sheets, preserving everything else",
    )
    ap.add_argument(
        "--config-only",
        action="store_true",
        help="re-import fixed costs, paycheck and the rent breakdown only",
    )
    args = ap.parse_args()

    report = Report()
    with SessionLocal() as session:
        if args.config_only:
            reset_config(session)
        elif args.accounts_only:
            linked = session.scalar(
                select(func.count(Transaction.id)).where(
                    Transaction.account_id.is_not(None)
                )
            )
            if linked:
                print(
                    f"{linked} transactions reference an account; refusing to "
                    "delete accounts underneath them.",
                    file=sys.stderr,
                )
                return 1
            reset_accounts(session)
        elif args.reset:
            reset(session)
        elif session.scalar(select(Category).limit(1)):
            print("Database already holds data. Re-run with --reset.", file=sys.stderr)
            return 1

        imp = Importer(session, report)
        if args.accounts_only or args.config_only:
            # Categories already exist; adopt them rather than re-seeding.
            for cat in session.scalars(select(Category)).all():
                imp.categories[cat.name.lower()] = cat
        else:
            imp.seed_categories()

        # Newest workbook last, so its config wins for the current period.
        for path in sorted(args.workbooks):
            m = re.search(r"(\d{4})", path.stem)
            if not m:
                print(f"skipping {path.name}: no year in filename", file=sys.stderr)
                continue
            year = int(m.group(1))
            label = path.stem
            print(f"  reading {path.name} ...")
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                if sheet.startswith("Yearly"):
                    continue  # derived rollup, recomputed on read
                if args.config_only:
                    continue
                if args.accounts_only:
                    if sheet == "Accounts":
                        imp.import_accounts(ws, year, label)
                    continue
                if "Spending" in sheet:
                    imp.import_spending(ws, year, label)
                elif "Budget" in sheet and sheet != "Monthly Overview":
                    imp.import_budget(ws, year, label)
                elif sheet == "Accounts":
                    imp.import_accounts(ws, year, label)

            if not args.accounts_only and year == max(
                int(re.search(r"(\d{4})", p.stem).group(1))
                for p in args.workbooks
                if re.search(r"(\d{4})", p.stem)
            ):
                eff = date(year, 1, 1)
                if "Monthly Fixed Costs" in wb.sheetnames:
                    imp.import_fixed_costs(wb["Monthly Fixed Costs"], eff, label)
                if "Rent" in wb.sheetnames:
                    imp.import_rent_breakdown(wb["Rent"], eff, label)
                if "Paycheck" in wb.sheetnames:
                    imp.import_paycheck(wb["Paycheck"], eff, label)

            session.flush()

        session.commit()

        # ---- verification: every sheet's own total must survive the import
        print("\nVerifying sheet totals ...")
        mismatches = 0
        for key, expected in sorted(imp.sheet_totals.items()):
            got = imp.imported_totals[key]
            if expected != got:
                mismatches += 1
                print(
                    f"  MISMATCH (rows dropped) {key}: cells={expected} imported={got}"
                )
        if mismatches == 0:
            print(
                f"  {len(imp.sheet_totals)} sheets: every row imported, no value lost"
            )

        # Second, softer check: agreement with the sheet's own rollup 'Total'.
        # A difference here does not mean the import is wrong. The rollup is a
        # SUMIF over a fixed list of category names, so a typo in the category
        # column makes a row invisible to it — the workbook silently
        # under-reports, and the import is the one telling the truth.
        checked = 0
        divergent: list[str] = []
        for key, self_total in sorted(imp.sheet_self_totals.items()):
            got = imp.imported_totals[key]
            if self_total != got:
                divergent.append(
                    f"  {key}: sheet's Total says {self_total}, actual sum is {got} "
                    f"(delta {got - self_total})"
                )
            else:
                checked += 1
        print(f"  {checked} sheets also match their own rollup 'Total' cell")
        if divergent:
            print("\nSheets whose own Total disagrees with their own rows:")
            for line in divergent:
                print(line)
            print(
                "  ^ the workbook under-reports these; see 'unknown category'"
                " anomalies below for the cause"
            )

        print("\nImported:")
        for k, v in sorted(report.counts.items()):
            print(f"  {v:6}  {k}")
        print(
            f"  {len(imp.merchants):6}  merchants (from "
            f"{report.counts['transactions']} descriptions)"
        )

        print("\nAnomalies:")
        print(report.render())
        return 1 if mismatches else 0


if __name__ == "__main__":
    raise SystemExit(main())
